"""Parsing domain services."""

from datetime import date
from datetime import datetime
from datetime import timedelta
from email.header import decode_header
from email.header import make_header
from email.message import Message
from fibaro_core.models import SettlementImport
from fibaro_core.services.summaries.periods import add_months
from fibaro_core.services.summaries.periods import month_label
from io import BytesIO
from pathlib import Path
from typing import Any
from typing import Dict
from typing import Iterable
from typing import Optional
from value_parsing import float_or_zero
import hashlib
import mimetypes
import re


def is_settlement_attachment(filename: str, content_type: str) -> bool:
    extension = Path(filename or "").suffix.lower()
    if extension in SETTLEMENT_ATTACHMENT_EXTENSIONS:
        return True
    return content_type in {
        "application/pdf",
        "image/jpeg",
        "image/png",
        "text/csv",
        "application/csv",
        "application/vnd.ms-excel",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "text/plain",
        "application/xml",
        "text/xml",
    }


def iter_message_attachments(message: Message) -> Iterable[Dict[str, Any]]:
    for part in message.walk():
        disposition = (part.get_content_disposition() or "").lower()
        filename = decoded_mime_header(part.get_filename())
        if disposition != "attachment" and not filename:
            continue
        payload = part.get_payload(decode=True)
        if not payload:
            continue
        content_type = part.get_content_type() or "application/octet-stream"
        if not filename:
            extension = mimetypes.guess_extension(content_type) or ".bin"
            filename = f"attachment{extension}"
        if not is_settlement_attachment(filename, content_type):
            continue
        yield {
            "filename": filename,
            "content_type": content_type,
            "bytes": payload,
            "sha256": hashlib.sha256(payload).hexdigest(),
        }


def parse_settlement_period(text_value: str, email_date: Optional[datetime]) -> tuple[Optional[date], Optional[date], str]:
    text_value = text_value or ""
    match = re.search(r"\b(20\d{2})[-_./ ](0?[1-9]|1[0-2])\b", text_value)
    if match:
        start = date(int(match.group(1)), int(match.group(2)), 1)
        end = add_months(start, 1) - timedelta(days=1)
        return start, end, month_label(start)
    match = re.search(r"\b(0?[1-9]|1[0-2])[-_./ ](20\d{2})\b", text_value)
    if match:
        start = date(int(match.group(2)), int(match.group(1)), 1)
        end = add_months(start, 1) - timedelta(days=1)
        return start, end, month_label(start)
    lowered = text_value.lower()
    for month_name, month_number in NORWEGIAN_MONTHS.items():
        match = re.search(rf"\b{re.escape(month_name)}\s+(20\d{{2}})\b", lowered)
        if match:
            start = date(int(match.group(1)), month_number, 1)
            end = add_months(start, 1) - timedelta(days=1)
            return start, end, month_label(start)
    if email_date and email_date.day <= 12:
        previous_month = add_months(email_date.date().replace(day=1), -1)
        return previous_month, add_months(previous_month, 1) - timedelta(days=1), f"{month_label(previous_month)} (antatt)"
    return None, None, "Ikke tolket"


def settlement_decode_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("latin-1", errors="replace")


def settlement_text_lines(text_value: str) -> list[str]:
    lines: list[str] = []
    for line in (text_value or "").replace("\u00a0", " ").splitlines():
        clean = re.sub(r"\s+", " ", line).strip()
        if clean:
            lines.append(clean)
    return lines


def extract_settlement_text(filename: str, content_type: str, content: bytes) -> Dict[str, Any]:
    extension = Path(filename or "").suffix.lower()
    method = "unknown"
    warnings: list[str] = []
    text_value = ""
    pages_count: Optional[int] = None
    try:
        if content_type == "application/pdf" or extension == ".pdf":
            method = "pypdf"
            try:
                from pypdf import PdfReader

                reader = PdfReader(BytesIO(content))
                pages_count = len(reader.pages)
                text_value = "\n".join(page.extract_text() or "" for page in reader.pages)
            except Exception as exc:
                warnings.append(f"PDF-tekst kunne ikke leses: {exc}")
        elif extension in {".xlsx", ".xls"} or content_type in {
            "application/vnd.ms-excel",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }:
            method = "openpyxl"
            try:
                from openpyxl import load_workbook

                workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
                rows: list[str] = []
                for sheet in workbook.worksheets:
                    rows.append(f"[Ark: {sheet.title}]")
                    for row in sheet.iter_rows(values_only=True):
                        values = [str(value).strip() for value in row if value not in (None, "")]
                        if values:
                            rows.append("\t".join(values))
                text_value = "\n".join(rows)
            except Exception as exc:
                warnings.append(f"Regneark kunne ikke leses: {exc}")
        else:
            method = "text"
            text_value = settlement_decode_text(content)
    except Exception as exc:
        warnings.append(f"Vedlegg kunne ikke tekstleses: {exc}")
    lines = settlement_text_lines(text_value)
    if not lines:
        warnings.append("Ingen tekst ble hentet fra vedlegget.")
    return {
        "method": method,
        "text": text_value,
        "lines": lines,
        "line_count": len(lines),
        "pages_count": pages_count,
        "warnings": warnings,
    }


def parse_settlement_number(value: Any) -> Optional[float]:
    if value is None:
        return None
    text_value = str(value).strip().replace("\u00a0", " ")
    if not text_value:
        return None
    text_value = text_value.replace(" ", "").replace("%", "")
    if "," in text_value and "." in text_value:
        if text_value.rfind(",") < text_value.rfind("."):
            text_value = text_value.replace(",", "")
        else:
            text_value = text_value.replace(".", "").replace(",", ".")
    elif "," in text_value:
        text_value = text_value.replace(",", ".")
    try:
        return float(text_value)
    except ValueError:
        return None


def settlement_numbers_from_line(line: str) -> list[float]:
    clean = (line or "").replace("\u00a0", " ").strip()
    if "," not in clean and "%" not in clean and re.fullmatch(r"-?[\d ]+", clean):
        tokens = clean.split()
        if len(tokens) >= 4 and len(tokens) % 2 == 0 and all(len(tokens[index + 1]) == 3 for index in range(0, len(tokens), 2)):
            values = []
            for index in range(0, len(tokens), 2):
                parsed = parse_settlement_number(f"{tokens[index]} {tokens[index + 1]}")
                if parsed is not None:
                    values.append(parsed)
            return values
    values: list[float] = []
    for match in SETTLEMENT_NUMBER_RE.findall(line or ""):
        parsed = parse_settlement_number(match)
        if parsed is not None:
            values.append(parsed)
    return values


def settlement_number_value(value: Optional[float]) -> Optional[float | int]:
    if value is None:
        return None
    if abs(value - round(value)) < 0.000001:
        return int(round(value))
    return round(value, 2)


def settlement_line_source(index: int, line: str) -> str:
    return f"PDF tekstlinje {index + 1}: {line}"


def settlement_parse_date_from_line(line: str) -> Optional[str]:
    match = re.search(r"\b(\d{1,2})\.(\d{1,2})\.(20\d{2})\b", line or "")
    if match:
        try:
            return date(int(match.group(3)), int(match.group(2)), int(match.group(1))).isoformat()
        except ValueError:
            return None
    match = re.search(r"\b(\d{1,2})/(\d{1,2})/(\d{2}|\d{4})\b", line or "")
    if match:
        year_value = int(match.group(3))
        if year_value < 100:
            year_value += 2000
        try:
            return date(year_value, int(match.group(2)), int(match.group(1))).isoformat()
        except ValueError:
            return None
    return None


def parse_parking_settlement_text(extraction: Dict[str, Any]) -> Dict[str, Any]:
    lines = list(extraction.get("lines") or [])
    parsed: Dict[str, Any] = {}
    field_sources: Dict[str, str] = {}
    field_confidence: Dict[str, float] = {}
    parser_notes: list[str] = []

    def set_field(field: str, value: Any, source: str, confidence: float = 0.9) -> None:
        if value is None or value == "":
            return
        parsed[field] = value
        field_sources[field] = source
        field_confidence[field] = round(max(0.0, min(1.0, confidence)), 2)

    for index, line in enumerate(lines):
        lower = line.lower()
        source = settlement_line_source(index, line)
        if lower.startswith("oslo "):
            set_field("report_date", settlement_parse_date_from_line(line), source, 0.95)
        period_match = re.search(r"oppgjørsrapport\s+for\s+(.+?20\d{2})\b", lower, flags=re.IGNORECASE)
        if period_match:
            label = period_match.group(1).strip()
            set_field("reported_period_label", label[:1].upper() + label[1:], source, 0.95)
        match = re.search(r"driftssted\s+(\d+)\s+(.+)", line, flags=re.IGNORECASE)
        if match:
            set_field("site_number", match.group(1), source, 0.95)
            set_field("site_name", match.group(2).strip(), source, 0.95)
        match = re.search(r"oppdragsgiver\s+(\d+)\s+(.+)", line, flags=re.IGNORECASE)
        if match:
            set_field("customer_number", match.group(1), source, 0.95)
            set_field("customer_name", match.group(2).strip(), source, 0.95)
        if lower.startswith("antall automater"):
            values = settlement_numbers_from_line(line)
            if values:
                set_field("machine_count", settlement_number_value(values[-1]), source, 0.95)
        if "mynt/kortautomat" in lower and "brutto" in lower:
            values = settlement_numbers_from_line(line)
            if values:
                set_field("gross_coin_card_ex_vat", settlement_number_value(values[-1]), source, 0.9)
        if lower.startswith("easypark"):
            values = settlement_numbers_from_line(line)
            if values:
                set_field("easypark_ex_vat", settlement_number_value(values[-1]), source, 0.95)
                set_field("easypark_inc_vat_estimate", settlement_number_value(values[-1] * 1.25), f"{source}; beregnet * 1,25", 0.75)
        if lower.startswith("fratrekk"):
            values = settlement_numbers_from_line(line)
            if values:
                set_field("settlement_fee_ex_vat", settlement_number_value(values[-1]), source, 0.9)
        email_matches = re.findall(r"[\w.\-+]+@[\w.\-]+\.\w+", line)
        if email_matches and "contact_email" not in parsed:
            set_field("contact_email", email_matches[0], source, 0.8)

    deduction_index = next((index for index, line in enumerate(lines) if line.lower().startswith("fratrekk")), None)
    numeric_rows: list[tuple[int, str, list[float]]] = []
    if deduction_index is not None:
        for index, line in enumerate(lines[deduction_index + 1 :], start=deduction_index + 1):
            lower = line.lower()
            if lower.startswith("betaling ") or lower.startswith("event.") or lower.startswith("nb!"):
                break
            values = settlement_numbers_from_line(line)
            if values and re.fullmatch(r"[-\d\s,%.]+", line):
                numeric_rows.append((index, line, values))

    assignments = [
        (
            ("revenue_basis_ex_vat", "revenue_share_percent", "revenue_share_ex_vat"),
            3,
            0.72,
            "Hovedgrunnlag/andel/sum lest fra første uetiketterte tallrad etter fratrekk.",
        ),
        (
            ("long_term_parking_ex_vat", "long_term_share_percent", "long_term_share_ex_vat"),
            3,
            0.68,
            "Langtidsparkering lest fra andre uetiketterte tallrad etter fratrekk.",
        ),
        (
            ("control_fee_net_ex_vat", "control_fee_share_percent", "control_fee_share_ex_vat"),
            3,
            0.68,
            "Kontrollavgift lest fra tredje uetiketterte tallrad etter fratrekk.",
        ),
        (
            ("total_basis_ex_vat", "total_share_ex_vat"),
            2,
            0.72,
            "Totalsum eks. mva lest fra fjerde uetiketterte tallrad etter fratrekk.",
        ),
        (
            ("vat_25_percent",),
            1,
            0.78,
            "25% mva lest fra tallraden etter totalsum.",
        ),
        (
            ("payout_inc_vat",),
            1,
            0.78,
            "Til utbetaling lest fra siste tallrad før betalingsinfo.",
        ),
    ]
    for row_index, assignment in enumerate(assignments):
        if row_index >= len(numeric_rows):
            parser_notes.append(f"Mangler tallrad for {', '.join(assignment[0])}.")
            continue
        fields, expected_count, confidence, note = assignment
        source_index, source_line, values = numeric_rows[row_index]
        if len(values) < expected_count:
            parser_notes.append(f"For få tall i tekstlinje {source_index + 1}: {source_line}")
            continue
        source = f"{settlement_line_source(source_index, source_line)}; {note}"
        for field, value in zip(fields, values):
            set_field(field, settlement_number_value(value), source, confidence)

    required_fields = [
        "reported_period_label",
        "site_number",
        "customer_number",
        "gross_coin_card_ex_vat",
        "easypark_ex_vat",
        "settlement_fee_ex_vat",
        "control_fee_net_ex_vat",
        "total_basis_ex_vat",
        "payout_inc_vat",
    ]
    found_required = sum(1 for field in required_fields if field in parsed)
    confidence = round(found_required / len(required_fields), 2)
    parsed["_meta"] = {
        "parser": "parking_parknordic_pdf",
        "parser_version": SETTLEMENT_PARSER_VERSION,
        "confidence": confidence,
        "field_sources": field_sources,
        "field_confidence": field_confidence,
        "line_count": extraction.get("line_count") or len(lines),
        "pages_count": extraction.get("pages_count"),
        "method": extraction.get("method"),
        "warnings": list(extraction.get("warnings") or []),
        "notes": parser_notes,
        "source_lines": lines,
    }
    return parsed


def parse_parking_settlement_attachment(filename: str, content_type: str, content: bytes) -> Dict[str, Any]:
    extraction = extract_settlement_text(filename, content_type, content)
    parsed = parse_parking_settlement_text(extraction)
    return parsed


def sun_settlement_number_from_line(line: str) -> Optional[float]:
    values = settlement_numbers_from_line(line)
    return settlement_number_value(values[-1]) if values else None


def normalize_sun_creditnote_signs(parsed: Dict[str, Any], field_sources: Dict[str, str]) -> bool:
    changed = False
    for field in SUN_SETTLEMENT_AMOUNT_FIELDS:
        value = parse_settlement_number(parsed.get(field))
        if value is None:
            continue
        normalized = settlement_number_value(-value)
        if normalized == parsed.get(field):
            continue
        parsed[field] = normalized
        field_sources[field] = f"{field_sources.get(field, 'Maskinlest fra oppgjørsskjema')}; fortegn snudd fordi dokumentet er kreditnota"
        changed = True
    return changed


def parse_sun_settlement_text(extraction: Dict[str, Any]) -> Dict[str, Any]:
    lines = list(extraction.get("lines") or [])
    parsed: Dict[str, Any] = {}
    field_sources: Dict[str, str] = {}
    field_confidence: Dict[str, float] = {}
    parser_notes: list[str] = []

    def set_field(field: str, value: Any, source: str, confidence: float = 0.9) -> None:
        if value is None or value == "":
            return
        parsed[field] = value
        field_sources[field] = source
        field_confidence[field] = round(max(0.0, min(1.0, confidence)), 2)

    line_fields = [
        ("solomsetning for perioden", "sun_revenue_ex_vat", 0.95),
        ("produktsalg for perioden", "product_sales_ex_vat", 0.95),
        ("transaksjonskostnad", "transaction_fee_ex_vat", 0.92),
        ("serviceavtale", "service_fee_ex_vat", 0.9),
        ("markedsf", "marketing_fee_ex_vat", 0.82),
        ("sum eks. mva", "sum_ex_vat", 0.95),
        ("mva grunnlag", "sum_ex_vat", 0.95),
        ("sum ordrelinjer", "sum_ex_vat", 0.9),
        ("25% mva", "vat_25_percent", 0.95),
        ("fakturabel", "payout_inc_vat", 0.95),
        ("belop nok", "payout_inc_vat", 0.95),
        ("beløp nok", "payout_inc_vat", 0.95),
    ]
    marketing_email_seen = False
    pending_field: Optional[tuple[str, float, str]] = None
    for index, line in enumerate(lines):
        lower = line.lower()
        source = settlement_line_source(index, line)
        if pending_field:
            values = settlement_numbers_from_line(line)
            if values:
                field, confidence, pending_source = pending_field
                set_field(field, settlement_number_value(values[-1]), f"{pending_source}; {source}", confidence)
                pending_field = None
            elif lower and not lower.endswith(":"):
                pending_field = None

        parsed_date = settlement_parse_date_from_line(line)
        if parsed_date and "credit_note_date" not in parsed:
            set_field("credit_note_date", parsed_date, source, 0.82)
            set_field("delivery_date", parsed_date, source, 0.72)
        if "kreditnotanr" in lower or "faktnr" in lower:
            values = settlement_numbers_from_line(line)
            if values:
                set_field("credit_note_number", settlement_number_value(values[-1]), source, 0.95)
        if "kreditnotadato" in lower:
            set_field("credit_note_date", parsed_date, source, 0.95)
        if "leveransedato" in lower:
            set_field("delivery_date", parsed_date, source, 0.95)
        if "organisasjonsnr" in lower or "foretaksregisteret" in lower:
            match = re.search(r"(NO\d+MVA|\d{9})", line, flags=re.IGNORECASE)
            if match:
                set_field("supplier_org_no", match.group(1), source, 0.9)
        if "altera as" in lower and "supplier_name" not in parsed:
            set_field("supplier_name", "Altera AS", source, 0.9)
        if "sun2 lillehammer" in lower and "customer_name" not in parsed:
            set_field("customer_name", line.strip(), source, 0.9)
        if lower.startswith("sum mva"):
            values = settlement_numbers_from_line(line)
            if values:
                set_field("vat_25_percent", settlement_number_value(values[-1]), source, 0.95)
            else:
                pending_field = ("vat_25_percent", 0.95, source)

        for marker, field, confidence in line_fields:
            if marker not in lower:
                continue
            value = sun_settlement_number_from_line(line)
            if field == "marketing_fee_ex_vat" and "e-post" in lower:
                field = "marketing_email_fee_ex_vat"
                marketing_email_seen = True
            elif field == "marketing_fee_ex_vat" and marketing_email_seen:
                field = "marketing_sms_fee_ex_vat"
            elif field == "marketing_fee_ex_vat":
                field = "marketing_sms_fee_ex_vat"
            set_field(field, value, source, confidence)

    is_creditnote = any("kreditnota" in line.lower() for line in lines)
    if is_creditnote and normalize_sun_creditnote_signs(parsed, field_sources):
        parser_notes.append("Fortegn er snudd fordi Altera sender oppgjøret som kreditnota. Inntekter vises positivt og fratrekk/gebyrer negativt.")

    sum_ex_vat = settlement_parsed_float(parsed, "sum_ex_vat")
    vat = settlement_parsed_float(parsed, "vat_25_percent")
    payout = settlement_parsed_float(parsed, "payout_inc_vat")
    if sum_ex_vat is not None and "vat_25_percent" not in parsed:
        set_field("vat_25_percent", settlement_number_value(sum_ex_vat * 0.25), "Beregnet fra Sum eks. MVA", 0.7)
    if sum_ex_vat is not None and vat is not None and payout is not None:
        diff = round(sum_ex_vat + vat - payout, 2)
        if abs(diff) > 1:
            parser_notes.append(f"Sum eks. mva + mva avviker fra Beløp NOK med {diff} kr.")

    if not lines:
        parser_notes.append("Vedlegget har ikke tekstlag. Originalen er lagret, men tall må kontrolleres manuelt eller OCR-leses senere.")

    required_fields = [
        "sun_revenue_ex_vat",
        "product_sales_ex_vat",
        "transaction_fee_ex_vat",
        "service_fee_ex_vat",
        "sum_ex_vat",
        "vat_25_percent",
        "payout_inc_vat",
    ]
    found_required = sum(1 for field in required_fields if field in parsed)
    confidence = round(found_required / len(required_fields), 2)
    parsed["_meta"] = {
        "parser": "sun_altera_creditnote_pdf",
        "parser_version": SETTLEMENT_PARSER_VERSION,
        "confidence": confidence,
        "field_sources": field_sources,
        "field_confidence": field_confidence,
        "line_count": extraction.get("line_count") or len(lines),
        "pages_count": extraction.get("pages_count"),
        "method": extraction.get("method"),
        "warnings": list(extraction.get("warnings") or []),
        "notes": parser_notes,
        "source_lines": lines,
    }
    return parsed


def parse_sun_settlement_attachment(filename: str, content_type: str, content: bytes) -> Dict[str, Any]:
    extraction = extract_settlement_text(filename, content_type, content)
    return parse_sun_settlement_text(extraction)


def parse_settlement_attachment_for_provider(provider: str, filename: str, content_type: str, content: bytes) -> Dict[str, Any]:
    if provider == SUN_SETTLEMENT_PROVIDER:
        return parse_sun_settlement_attachment(filename, content_type, content)
    return parse_parking_settlement_attachment(filename, content_type, content)


def settlement_period_from_parsed_dates(parsed: Any, *fields: str) -> tuple[Optional[date], Optional[date], str]:
    for field in fields:
        raw_value = settlement_parsed_value(parsed, field)
        if not raw_value:
            continue
        try:
            parsed_date = date.fromisoformat(str(raw_value))
        except ValueError:
            continue
        start = parsed_date.replace(day=1)
        return start, add_months(start, 1) - timedelta(days=1), month_label(start)
    return None, None, "Ikke tolket"


def settlement_public_parsed(parsed: Any) -> Dict[str, Any]:
    if not isinstance(parsed, dict):
        return {}
    return {key: value for key, value in parsed.items() if not str(key).startswith("_")}


def settlement_parsed_meta(parsed: Any) -> Dict[str, Any]:
    if not isinstance(parsed, dict):
        return {}
    meta = parsed.get("_meta")
    return meta if isinstance(meta, dict) else {}


def settlement_parsed_value(parsed: Any, field: str) -> Any:
    return settlement_public_parsed(parsed).get(field)


def settlement_parsed_float(parsed: Any, field: str) -> Optional[float]:
    value = settlement_parsed_value(parsed, field)
    if isinstance(value, (int, float)):
        return float(value)
    return parse_settlement_number(value)


def settlement_needs_parse(row: SettlementImport) -> bool:
    meta = settlement_parsed_meta(row.parsed)
    return meta.get("parser_version") != SETTLEMENT_PARSER_VERSION


def ensure_settlement_parsed(row: SettlementImport) -> bool:
    if not settlement_needs_parse(row):
        return False
    parsed = parse_settlement_attachment_for_provider(
        row.provider,
        row.attachment_filename or "",
        row.attachment_content_type or "",
        row.attachment_bytes or b"",
    )
    row.parsed = parsed
    raw = dict(row.raw or {}) if isinstance(row.raw, dict) else {}
    raw["settlement_parser"] = settlement_parsed_meta(parsed)
    row.raw = raw
    if not row.period_start and parsed.get("reported_period_label"):
        period_start, period_end, period_label = parse_settlement_period(str(parsed.get("reported_period_label")), row.email_date)
        row.period_start = period_start
        row.period_end = period_end
        row.period_label = period_label
    if not row.period_start and row.provider == SUN_SETTLEMENT_PROVIDER:
        period_start, period_end, period_label = settlement_period_from_parsed_dates(parsed, "delivery_date", "credit_note_date")
        row.period_start = period_start
        row.period_end = period_end
        row.period_label = period_label
    meta = settlement_parsed_meta(parsed)
    row.status = "tolket" if float_or_zero(meta.get("confidence")) >= 0.6 else "krever kontroll"
    return True


def settlement_field_source(parsed: Any, field: str, fallback: str = "Maskinlest fra oppgjørsskjema") -> str:
    sources = settlement_parsed_meta(parsed).get("field_sources")
    if isinstance(sources, dict):
        return str(sources.get(field) or fallback)
    return fallback


def settlement_field_confidence(parsed: Any, field: str) -> Optional[float]:
    values = settlement_parsed_meta(parsed).get("field_confidence")
    if not isinstance(values, dict):
        return None
    raw_value = values.get(field)
    if isinstance(raw_value, (int, float)):
        return float(raw_value)
    return parse_settlement_number(raw_value)


PARKING_SETTLEMENT_PROVIDER = "parking_parknordic"


SUN_SETTLEMENT_PROVIDER = "sun_altera"


SETTLEMENT_ATTACHMENT_EXTENSIONS = {".pdf", ".csv", ".xlsx", ".xls", ".xml", ".txt", ".jpg", ".jpeg", ".png"}


SETTLEMENT_PARSER_VERSION = 5


NORWEGIAN_MONTHS = {
    "januar": 1,
    "jan": 1,
    "februar": 2,
    "feb": 2,
    "mars": 3,
    "mar": 3,
    "april": 4,
    "apr": 4,
    "mai": 5,
    "juni": 6,
    "jun": 6,
    "juli": 7,
    "jul": 7,
    "august": 8,
    "aug": 8,
    "september": 9,
    "sep": 9,
    "oktober": 10,
    "okt": 10,
    "november": 11,
    "nov": 11,
    "desember": 12,
    "des": 12,
}


SETTLEMENT_NUMBER_RE = re.compile(
    r"-?(?:\d{1,3}(?:[ \u00a0]\d{3})+(?:[,.]\d+)?|\d{1,3}(?:,\d{3})+(?:\.\d+)?|\d+\.\d+|\d+,\d+|\d+)"
)


SUN_SETTLEMENT_AMOUNT_FIELDS = {
    "sun_revenue_ex_vat",
    "product_sales_ex_vat",
    "transaction_fee_ex_vat",
    "service_fee_ex_vat",
    "marketing_sms_fee_ex_vat",
    "marketing_email_fee_ex_vat",
    "sum_ex_vat",
    "vat_25_percent",
    "payout_inc_vat",
}


def decoded_mime_header(value: Optional[str]) -> str:
    if not value:
        return ""
    try:
        return str(make_header(decode_header(value)))
    except Exception:
        return value
